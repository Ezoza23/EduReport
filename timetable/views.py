from datetime import date, timedelta
import calendar
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from .models import Teacher, TimetableCard, LessonRecord, MonthlyReport, ScheduledLesson, Subject, Classroom, DepartmentReport, ExamAttendance
import holidays as holidays_lib

def get_uz_holidays(year, month):
    from timetable.models import RedDay
    import holidays as holidays_lib

    uz_holidays = holidays_lib.Uzbekistan(years=year)
    result = {}

    # Official holidays
    for d, name in uz_holidays.items():
        if d.month == month:
            result[str(d)] = name

    # Custom red days from admin
    custom_red_days = RedDay.objects.filter(date__year=year, date__month=month)
    for rd in custom_red_days:
        result[str(rd.date)] = rd.reason

    return result

TRANSLATIONS = {
    "en": {
        # ... existing keys ...
        "months": ["January", "February", "March", "April", "May", "June",
                   "July", "August", "September", "October", "November", "December"],
        "days_short": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"],
        "days_full": ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"],
        # report headers
        "report_date": "Date",
        "report_day": "Day",
        "report_subject": "Subject",
        "report_period": "Period",
        "report_room": "Room",
        "report_status": "Status",
        "report_note": "Note",
        "report_covered": "Covered",
        "report_missed": "Missed",
        "report_pending": "Pending",
        "report_rescheduled": "Rescheduled",
        "report_not_covered": "Not Covered",
        "report_total": "Total",
        "logout": "Logout",
        "dean_dashboard": "Dean Dashboard",
"report_type": "Type",
    "type_lecture": "Lecture",
    "type_seminar": "Seminar",
    "type_lab": "Lab",
    "type_exam": "Exam",
"type_final_exam": "Final Exam",
"type_midterm_exam": "Midterm Exam",
"type_retake_exam": "Retake Exam",
"manually_added": "Added by teacher",
    },
    "uz": {
        # ... existing keys ...
        "months": ["Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
                   "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"],
        "days_short": ["Du", "Se", "Ch", "Pa", "Ju", "Sh", "Ya"],
        "days_full": ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"],
        # report headers
        "report_date": "Sana",
        "report_day": "Kun",
        "report_subject": "Fan",
        "report_period": "Dars",
        "report_room": "Xona",
        "report_status": "Holat",
        "report_note": "Izoh",
        "report_covered": "O'tildi",
        "report_missed": "O'tilmadi",
        "report_pending": "Kutilmoqda",
        "report_rescheduled": "Ko'chirildi",
        "report_not_covered": "O'tilmadi",
        "report_total": "Jami",
        "logout": "Chiqish",
    "dean_dashboard": "Dekan paneli",
"report_type": "Tur",
    "type_lecture": "Ma'ruza",
    "type_seminar": "Seminar",
    "type_lab": "Laboratoriya",
    "type_final_exam": "Yakuniy imtihon",
"type_midterm_exam": "Oraliq nazorat",
"type_retake_exam": "Qayta topshirish",
"manually_added": "O'qituvchi tomonidan qo'shilgan",
    }
}
PERIOD_TIMES = {
    "1": ("09:00", "10:10"),
    "2": ("10:20", "11:30"),
    "3": ("11:40", "12:50"),
    "4": ("13:30", "14:40"),
    "5": ("14:50", "16:00"),
    "6": ("16:10", "17:20"),
    "7": ("17:30", "18:40"),
    "8": ("18:50", "20:00"),
}

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
PERIODS = ["1", "2", "3", "4", "5", "6", "7", "8"]


def get_lesson_type(classroom_name):
    if classroom_name and "lecture" in classroom_name.lower():
        return "lecture"
    return "seminar"


def login_view(request):
    error = None
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            if user.is_superuser:
                # Sync teachers from EduPage on admin login
                try:
                    from django.core.management import call_command
                    call_command("sync_edupage")
                except Exception as e:
                    print(f"Admin sync failed: {e}")
                return redirect("admin_panel")
            teacher = Teacher.objects.filter(user=user).first()
            if teacher:
                try:
                    from .sync_teacher import sync_teacher_schedule
                    created, updated = sync_teacher_schedule(teacher)
                    print(f"Login sync: {created} created, {updated} updated for {teacher.full_name}")
                except Exception as e:
                    print(f"Login sync failed: {e}")
            if teacher and teacher.is_dean:
                return redirect("dean_dashboard")
            return redirect("timetable")
        else:
            error = "Invalid username or password"
    return render(request, "login.html", {"error": error})


def logout_view(request):
    logout(request)
    return redirect("login")

from django.views.decorators.csrf import ensure_csrf_cookie

@ensure_csrf_cookie
@login_required(login_url="login")
def timetable_view(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    if not teacher:
        return render(request, "timetable.html", {"error": "No timetable linked to your account"})

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    num_days = calendar.monthrange(year, month)[1]
    month_days = [date(year, month, d) for d in range(1, num_days + 1)]
    month_days = [d for d in month_days if d.weekday() < 6]

    # Group into weeks
    weeks = []
    current_week = []
    for d in month_days:
        if d.weekday() == 0 and current_week:
            weeks.append(current_week)
            current_week = []
        current_week.append(d)
    if current_week:
        weeks.append(current_week)

    period_list = get_period_list()
    PERIODS_ACTIVE = get_periods_list_numbers()

    # Get scheduled lessons for this month
    scheduled_lessons = ScheduledLesson.objects.filter(
        teacher=teacher,
        date__year=year,
        date__month=month
    ).select_related("subject", "classroom")

    scheduled_map = {(str(sl.date), sl.period): sl for sl in scheduled_lessons}

    # Get lesson records
    records = LessonRecord.objects.filter(
        teacher=teacher,
        date__year=year,
        date__month=month
    )
    record_map = {(r.scheduled_lesson_id, str(r.date)): r for r in records}


    can_submit = True
    report = MonthlyReport.objects.filter(teacher=teacher, year=year, month=month).order_by("-id").first()
    already_submitted = report and report.status == "submitted"
    is_rejected = bool(
        report and report.status == "submitted" and (
                report.dean_approval == "rejected" or report.admin_approval == "rejected"
        )
    )
    edit_mode = request.GET.get("edit") == "1"
    can_edit = (not already_submitted) or (is_rejected and edit_mode)
    lang = request.session.get("lang", "en")
    tr = TRANSLATIONS[lang]
    weeks_data = []
    holidays = get_uz_holidays(year, month)
    for week in weeks:
        rows = []
        for d in week:
            day_name = DAYS[d.weekday()]
            day_short = tr["days_short"][d.weekday()]
            cells = []
            for period in PERIODS_ACTIVE:
                sl = scheduled_map.get((str(d), period))
                record = record_map.get((sl.id if sl else None, str(d)))

                # Freeze past days — auto snapshot
                if sl and d < today and not record:
                    record, _ = LessonRecord.objects.get_or_create(
                        scheduled_lesson=sl,
                        date=d,
                        defaults={
                            "teacher": teacher,
                            "is_covered": None,
                            "is_replaced": False,
                        }
                    )
                    record_map[(sl.id, str(d))] = record

                    # Hide original slot if it has been rescheduled to another date
                is_rescheduled_away = record and record.is_replaced and record.replacement_date and record.replacement_date != d

                cells.append({
                    "lesson": sl if not is_rescheduled_away else None,
                    "record": record,
                    "date": str(d),
                    "period": period,
                    "is_replacement": sl.is_rescheduled_slot if sl else False,
                    "original_lesson_id": sl.rescheduled_from_id if (sl and sl.is_rescheduled_slot) else None,
                    "original_lesson_date": str(sl.rescheduled_from.date) if (
                                sl and sl.is_rescheduled_slot and sl.rescheduled_from) else None,
                    "is_custom": sl.is_custom_added if sl else False,
                })
            rows.append({
                "date": d,
                "day_name": day_name,
                "day_short": day_short,
                "cells": cells,
            })
        weeks_data.append(rows)

    prev_month = date(year, month, 1) - timedelta(days=1)
    next_month = date(year, month, num_days) + timedelta(days=1)
    lang = request.session.get("lang", "en")
    tr = TRANSLATIONS[lang]
    exam_attendance = ExamAttendance.objects.filter(
        teacher=teacher, year=year, month=month
    ).first()
    return render(request, "timetable.html", {
        "teacher": teacher,
        "period_list": period_list,
        "weeks_data": weeks_data,
        "holidays": holidays,
        "month_name": f"{tr['months'][month - 1]} {year}",
        "year": year,
        "month": month,
        "prev": {"year": prev_month.year, "month": prev_month.month},
        "next": {"year": next_month.year, "month": next_month.month},
        "can_submit": can_submit,
        "already_submitted": already_submitted,
        "can_edit": can_edit,
        "report": report,
        "lang": lang,
        "tr": tr,
        "exam_attendance": exam_attendance,
        "is_rejected": is_rejected,
        "edit_mode": edit_mode,
    })

@login_required(login_url="login")
def dean_dashboard(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    if not teacher or not teacher.is_dean:
        return redirect("timetable")

    lang = request.session.get("lang", "en")
    tr = TRANSLATIONS[lang]
    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))
    holidays = get_uz_holidays(year, month)
    num_days = calendar.monthrange(year, month)[1]
    all_days = [date(year, month, d) for d in range(1, num_days + 1)]
    working_days = [d for d in all_days if d.weekday() < 6]

    dept_teachers = Teacher.objects.filter(
        department=teacher.department,
        user__isnull=False
    ).order_by("full_name")

    teacher_rows = []
    for t in dept_teachers:
        records = LessonRecord.objects.filter(
            teacher=t,
            date__year=year,
            date__month=month,
            is_covered=True
        )
        day_counts = {}
        for d in working_days:
            count = records.filter(date=d).count()
            day_counts[str(d)] = count

        total = records.count()
        report = MonthlyReport.objects.filter(teacher=t, year=year, month=month).order_by("-id").first()
        ea = ExamAttendance.objects.filter(teacher=t, year=year, month=month).first()
        teacher_rows.append({
            "teacher": t,
            "day_counts": day_counts,
            "total": total,
            "report": report,
            "final_exam_count": ea.final_exam_count if ea else 0,
        "final_exam_unit": ea.final_exam_unit if ea else "students",
        "midterm_exam_count": ea.midterm_exam_count if ea else 0,
        "midterm_exam_unit": ea.midterm_exam_unit if ea else "students",
        "retake_exam_count": ea.retake_exam_count if ea else 0,
        "retake_exam_unit": ea.retake_exam_unit if ea else "students",
        })

    dept_report = None
    if teacher.department:
        dept_report = DepartmentReport.objects.filter(
            department=teacher.department, year=year, month=month
        ).first()

    prev_month = date(year, month, 1) - timedelta(days=1)
    next_month_date = date(year, month, num_days) + timedelta(days=1)

    return render(request, "dean_dashboard.html", {
        "teacher": teacher,
        "teacher_rows": teacher_rows,
        "working_days": working_days,
        "month_name": f"{tr['months'][month - 1]} {year}",
        "year": year,
        "month": month,
        "prev": {"year": prev_month.year, "month": prev_month.month},
        "next": {"year": next_month_date.year, "month": next_month_date.month},
        "lang": lang,
        "tr": tr,
        "holidays": holidays,
        "holidays_json": holidays,
        "dept_report": dept_report,
        "week_spans": get_week_spans(working_days),
    })

@login_required(login_url="login")
@require_POST
def mark_lesson(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    lesson_id = request.POST.get("lesson_id")
    date_str = request.POST.get("date")
    is_covered = request.POST.get("is_covered")
    note = request.POST.get("note", "")
    replacement_date = request.POST.get("replacement_date", "")

    sl = get_object_or_404(ScheduledLesson, id=lesson_id, teacher=teacher)
    lesson_date = date.fromisoformat(date_str)

    existing = LessonRecord.objects.filter(
        scheduled_lesson=sl,
        date=lesson_date
    ).first()

    if replacement_date:
        rep_date = date.fromisoformat(replacement_date)

        # Find the true original lesson first
        if sl.is_rescheduled_slot and sl.rescheduled_from:
            true_original = sl.rescheduled_from
        else:
            true_original = sl

        # Save original data before any deletion
        original_subject = true_original.subject
        original_classroom = true_original.classroom
        original_class_names = true_original.class_names
        original_groups = list(true_original.groups.all())
        original_date = true_original.date
        original_period = true_original.period

        # Get teacher-selected period
        replacement_period = request.POST.get("replacement_period", "").strip()
        period = replacement_period if replacement_period else original_period

        # Check if returning to exact original slot
        is_same_slot = (rep_date == original_date) and (period == original_period)

        # Get existing record on the original lesson
        existing_record = LessonRecord.objects.filter(
            scheduled_lesson=true_original,
            date=original_date,
        ).first()

        # Clean up ALL existing replacement slots for this original
        ScheduledLesson.objects.filter(
            teacher=teacher,
            rescheduled_from=true_original,
            is_rescheduled_slot=True,
        ).delete()

        # Update the original lesson's record
        record, _ = LessonRecord.objects.update_or_create(
            scheduled_lesson=true_original,
            date=original_date,
            defaults={
                "teacher": teacher,
                "is_covered": existing_record.is_covered if existing_record else False,
                "note": note,
                "is_replaced": not is_same_slot,
                "replacement_date": rep_date if not is_same_slot else None,
            }
        )

        if not is_same_slot:
            rep_day = DAYS[rep_date.weekday()]

            # Create new replacement slot
            rep_sl, created = ScheduledLesson.objects.get_or_create(
                teacher=teacher,
                date=rep_date,
                period=period,
                defaults={
                    "subject": original_subject,
                    "classroom": original_classroom,
                    "class_names": original_class_names,
                    "day": rep_day,
                    "is_rescheduled_slot": True,
                    "rescheduled_from": true_original,
                }
            )
            if created:
                rep_sl.groups.set(original_groups)
            else:
                # Slot already taken — update it to be our replacement
                rep_sl.is_rescheduled_slot = True
                rep_sl.rescheduled_from = true_original
                rep_sl.save()
                rep_sl.groups.set(original_groups)

    else:
        # Just toggling covered status
        record, _ = LessonRecord.objects.update_or_create(
            scheduled_lesson=sl,
            date=lesson_date,
            defaults={
                "teacher": teacher,
                "is_covered": is_covered == "true",
                "note": note,
                "is_replaced": existing.is_replaced if existing else False,
                "replacement_date": existing.replacement_date if existing else None,
            }
        )

    return JsonResponse({
        "status": "ok",
        "is_covered": record.is_covered,
        "is_replaced": record.is_replaced
    })

@login_required(login_url="login")
@require_POST
def submit_report(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    year = int(request.POST.get("year"))
    month = int(request.POST.get("month"))

    latest = MonthlyReport.objects.filter(teacher=teacher, year=year, month=month).order_by("-id").first()

    if latest is None or latest.dean_approval == "rejected" or latest.admin_approval == "rejected":
        MonthlyReport.objects.create(
            teacher=teacher, year=year, month=month,
            status="submitted", submitted_at=timezone.now(),
            dean_approval="pending", admin_approval="waiting_dean",
        )
    else:
        latest.status = "submitted"
        latest.submitted_at = timezone.now()
        latest.dean_approval = "pending"
        latest.admin_approval = "waiting_dean"
        latest.save()

    return JsonResponse({"status": "submitted"})

@login_required(login_url="login")
@require_POST
def cancel_report(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    year = int(request.POST.get("year"))
    month = int(request.POST.get("month"))

    latest = MonthlyReport.objects.filter(teacher=teacher, year=year, month=month).order_by("-id").first()
    if latest and latest.status == "submitted" and latest.dean_approval == "pending" and latest.admin_approval == "waiting_dean":
        latest.status = "draft"
        latest.submitted_at = None
        latest.save()

    try:
        from .sync_teacher import sync_teacher_schedule
        created, updated = sync_teacher_schedule(teacher)
    except Exception as e:
        print(f"Cancel-sync failed: {e}")

    return JsonResponse({"status": "cancelled"})






import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


@login_required(login_url="login")
def export_dean_excel(request):
    requesting_teacher = Teacher.objects.filter(user=request.user).first()

    if request.user.is_superuser:
        dept_id = request.GET.get("department_id")
        if not dept_id:
            return redirect("admin_departments")
        department = get_object_or_404(Department, id=dept_id)
    else:
        if not requesting_teacher or not requesting_teacher.is_dean:
            return redirect("timetable")
        department = requesting_teacher.department
        if not department:
            return redirect("timetable")

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    num_days = calendar.monthrange(year, month)[1]
    all_days = [date(year, month, d) for d in range(1, num_days + 1)]
    working_days = [d for d in all_days if d.weekday() < 6]

    dept_teachers = Teacher.objects.filter(
        department=department,
        user__isnull=False
    ).order_by("full_name")

    teacher_rows = []
    for t in dept_teachers:
        records = LessonRecord.objects.filter(
            teacher=t,
            date__year=year,
            date__month=month,
            is_covered=True
        )
        day_counts = {str(d): records.filter(date=d).count() for d in working_days}
        total = records.count()
        teacher_rows.append({
            "name": t.full_name,
            "day_counts": day_counts,
            "total": total,
            "final_exam_count": t.final_exam_count,
            "midterm_exam_count": t.midterm_exam_count,
            "retake_exam_count": t.retake_exam_count,
            "final_exam_unit": t.final_exam_unit,
            "midterm_exam_unit": t.midterm_exam_unit,
            "retake_exam_unit": t.retake_exam_unit,
        })

    wb = Workbook()
    ws = wb.active
    month_name = date(year, month, 1).strftime("%B %Y")
    dept_name = department.name if department else "Department"
    ws.title = month_name

    header_fill = PatternFill("solid", start_color="1B497D", end_color="1B497D")
    exam_header_fill = PatternFill("solid", start_color="2E5F94", end_color="2E5F94")
    total_fill = PatternFill("solid", start_color="163D6A", end_color="163D6A")
    has_lesson_fill = PatternFill("solid", start_color="E8F5EC", end_color="E8F5EC")
    exam_count_fill = PatternFill("solid", start_color="EEF4FF", end_color="EEF4FF")
    white_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    normal_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", bold=True, size=10)
    blue_bold_font = Font(name="Arial", bold=True, color="1B497D", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    final_col = len(working_days) + 2
    midterm_col = final_col + 1
    retake_col = midterm_col + 1
    total_col = retake_col + 1

    last_col_letter = get_column_letter(total_col)
    ws.merge_cells(f"A1:{last_col_letter}1")
    title_cell = ws["A1"]
    title_cell.value = f"{dept_name} — {month_name}"
    title_cell.font = Font(name="Arial", bold=True, size=14, color="1B497D")
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    ws.row_dimensions[2].height = 36
    ws["A2"].value = "Teacher"
    ws["A2"].font = white_font
    ws["A2"].fill = header_fill
    ws["A2"].alignment = left
    ws["A2"].border = border
    ws.column_dimensions["A"].width = 28

    for i, d in enumerate(working_days):
        col = i + 2
        col_letter = get_column_letter(col)
        cell = ws.cell(row=2, column=col)
        cell.value = f"{d.day}\n{d.strftime('%a')}"
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[col_letter].width = 6

    lang = request.session.get("lang", "en")
    tr = TRANSLATIONS[lang]
    exam_headers = [
        (final_col, tr.get("final_exam", "Final Exam")),
        (midterm_col, tr.get("midterm_exam", "Midterm Exam")),
        (retake_col, tr.get("retake_exam", "Retake Exam")),
    ]
    for col, label in exam_headers:
        col_letter = get_column_letter(col)
        cell = ws.cell(row=2, column=col)
        cell.value = label
        cell.font = white_font
        cell.fill = exam_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[col_letter].width = 11

    total_col_letter = get_column_letter(total_col)
    total_cell = ws.cell(row=2, column=total_col)
    total_cell.value = "Total"
    total_cell.font = white_font
    total_cell.fill = total_fill
    total_cell.alignment = center
    total_cell.border = border
    ws.column_dimensions[total_col_letter].width = 8

    for row_idx, row in enumerate(teacher_rows):
        excel_row = row_idx + 3
        ws.row_dimensions[excel_row].height = 22

        name_cell = ws.cell(row=excel_row, column=1)
        name_cell.value = row["name"]
        name_cell.font = bold_font
        name_cell.alignment = left
        name_cell.border = border
        name_cell.fill = PatternFill("solid", start_color="F7F9FC", end_color="F7F9FC")

        for i, d in enumerate(working_days):
            col = i + 2
            count = row["day_counts"].get(str(d), 0)
            cell = ws.cell(row=excel_row, column=col)
            cell.value = count if count else None
            cell.font = bold_font if count else normal_font
            cell.alignment = center
            cell.border = border
            if count:
                cell.fill = has_lesson_fill

        unit_map = {"final_exam_count": "final_exam_unit", "midterm_exam_count": "midterm_exam_unit",
                    "retake_exam_count": "retake_exam_unit"}
        for col, key in [(final_col, "final_exam_count"), (midterm_col, "midterm_exam_count"),
                         (retake_col, "retake_exam_count")]:
            value = row[key]
            unit = row.get(unit_map[key], "students")
            cell = ws.cell(row=excel_row, column=col)
            cell.value = f"{value} {'groups' if unit == 'groups' else 'students'}" if value else None
            cell.font = bold_font if value else normal_font
            cell.alignment = center
            cell.border = border
            if value:
                cell.fill = exam_count_fill

        total_cell = ws.cell(row=excel_row, column=total_col)
        total_cell.value = row["total"]
        total_cell.font = blue_bold_font
        total_cell.alignment = center
        total_cell.border = border
        total_cell.fill = PatternFill("solid", start_color="EEF4FF", end_color="EEF4FF")

    filename = f"{dept_name}_{month_name}.xlsx".replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.read())
    return response

@login_required(login_url="login")
def download_report(request, teacher_id=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    lang = request.session.get("lang", "en")
    tr = TRANSLATIONS[lang]

    if teacher_id:
        if not request.user.is_superuser:
            requesting_teacher = Teacher.objects.filter(user=request.user).first()
            if not requesting_teacher or not requesting_teacher.is_dean:
                return redirect("timetable")
        teacher = get_object_or_404(Teacher, id=teacher_id)
    else:
        teacher = Teacher.objects.filter(user=request.user).first()
        if not teacher:
            return redirect("timetable")

    # ... everything below this point is unchanged from before ...

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))
    month_name = f"{tr['months'][month - 1]} {year}"

    lessons = ScheduledLesson.objects.filter(
        teacher=teacher,
        date__year=year,
        date__month=month,
    ).select_related("subject", "classroom").order_by("date", "period")

    records = LessonRecord.objects.filter(
        teacher=teacher,
        date__year=year,
        date__month=month,
    )
    record_map = {r.scheduled_lesson_id: r for r in records}
    report = MonthlyReport.objects.filter(teacher=teacher, year=year, month=month).order_by("-id").first()

    wb = Workbook()
    ws = wb.active
    ws.title = month_name[:31]

    # Styles
    header_fill = PatternFill("solid", start_color="1B497D", end_color="1B497D")
    covered_fill = PatternFill("solid", start_color="E8F5EC", end_color="E8F5EC")
    missed_fill = PatternFill("solid", start_color="FDECEA", end_color="FDECEA")
    rescheduled_fill = PatternFill("solid", start_color="FFF8E1", end_color="FFF8E1")
    white_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    normal_font = Font(name="Arial", size=10)
    covered_font = Font(name="Arial", bold=True, color="155724", size=10)
    missed_font = Font(name="Arial", bold=True, color="721C24", size=10)
    rescheduled_font = Font(name="Arial", bold=True, color="856404", size=10)
    pending_font = Font(name="Arial", size=10, color="888888")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = f"{teacher.full_name} — {month_name}"
    title.font = Font(name="Arial", bold=True, size=14, color="1B497D")
    title.alignment = center
    ws.row_dimensions[1].height = 32

    # Summary counters — filled after the data loop
    total = lessons.count()
    covered_count = 0
    rescheduled_count = 0
    missed_count = 0
    ws.row_dimensions[2].height = 40  # taller — summary now has 2 lines

    # Headers
    # Headers
    headers = [
        tr["report_date"],
        tr["report_day"],
        tr["report_subject"],
        tr["report_period"],
        tr["report_room"],
        tr.get("report_type", "Type"),
        tr["report_status"],
        tr["report_note"],
    ]
    col_widths = [14, 14, 35, 10, 25, 15, 16, 45]  # 8 widths now, matching 8 headers

    ws.row_dimensions[3].height = 28
    for i, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=i)
        cell.value = header
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = width

    # Data rows
    for row_idx, lesson in enumerate(lessons, 4):
        ws.row_dimensions[row_idx].height = 22
        record = record_map.get(lesson.id)
        is_submitted = report is not None and report.status == "submitted"

        if record and record.is_replaced:
            # Rescheduled lesson
            if record.is_covered:
                status = tr["report_covered"]
                row_fill = covered_fill
                status_font = covered_font
            elif is_submitted:
                status = tr["report_missed"]
                row_fill = missed_fill
                status_font = missed_font
            else:
                status = tr["report_pending"]
                row_fill = rescheduled_fill
                status_font = rescheduled_font

            # Build note with replacement date and period
            if record.replacement_date:
                rep_date_str = record.replacement_date.strftime("%d %b %Y")
                rep_slot = ScheduledLesson.objects.filter(
                    teacher=teacher,
                    date=record.replacement_date,
                    is_rescheduled_slot=True,
                ).first()
                if rep_slot:
                    note = f"Rescheduled → {rep_date_str} {tr['report_period']} {rep_slot.period}"
                else:
                    note = f"Rescheduled → {rep_date_str}"
                if record.note:
                    note += f" — {record.note}"
            else:
                note = "Rescheduled"
                if record.note:
                    note += f" — {record.note}"

            rescheduled_count += 1
            if status == tr["report_covered"]:
                covered_count += 1
            elif status == tr["report_missed"]:
                missed_count += 1

        elif record and record.is_covered:
            status = tr["report_covered"]
            row_fill = covered_fill
            status_font = covered_font
            note = record.note or "—"
            covered_count += 1

        elif record and record.is_covered == False:
            if is_submitted:
                status = tr["report_missed"]
                row_fill = missed_fill
                status_font = missed_font
                missed_count += 1
            else:
                status = tr["report_pending"]
                row_fill = None
                status_font = pending_font
            note = record.note or "—"

        else:
            if is_submitted:
                status = tr["report_missed"]
                row_fill = missed_fill
                status_font = missed_font
                missed_count += 1
            else:
                status = tr["report_pending"]
                row_fill = None
                status_font = pending_font
            note = "—"
        if lesson.is_custom_added:
            marker = tr.get("manually_added", "Added by teacher")
            if note and note != "—":
                note = f"{marker} — {note}"
            else:
                note = marker
        type_labels = {
            "lecture": tr.get("type_lecture", "Lecture"),
            "seminar": tr.get("type_seminar", "Seminar"),
            "lab": tr.get("type_lab", "Lab"),
            "exam": tr.get("type_exam", "Exam"),
            "final_exam": tr.get("type_final_exam", "Final Exam"),
            "midterm_exam": tr.get("type_midterm_exam", "Midterm Exam"),
            "retake_exam": tr.get("type_retake_exam", "Retake Exam"),
        }
        data = [
            lesson.date.strftime("%d %b %Y"),
            tr["days_full"][lesson.date.weekday()],
            str(lesson.subject) if lesson.subject else "—",
            f"{tr['report_period']} {lesson.period}",
            str(lesson.classroom) if lesson.classroom else "—",
            type_labels.get(lesson.lesson_type, lesson.lesson_type.capitalize()),
            status,
            note,
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = left if col_idx in [3, 7] else center
            if row_fill:
                cell.fill = row_fill
            if col_idx == 6:
                cell.font = status_font
            else:
                cell.font = normal_font
    exam_attendance = ExamAttendance.objects.filter(
        teacher=teacher, year=year, month=month
    ).first()
    # Write summary after loop completes
    ws.merge_cells("A2:G2")
    summary = ws["A2"]
    final_count = exam_attendance.final_exam_count if exam_attendance else 0
    final_unit = exam_attendance.final_exam_unit if exam_attendance else "students"
    midterm_count = exam_attendance.midterm_exam_count if exam_attendance else 0
    midterm_unit = exam_attendance.midterm_exam_unit if exam_attendance else "students"
    retake_count = exam_attendance.retake_exam_count if exam_attendance else 0
    retake_unit = exam_attendance.retake_exam_unit if exam_attendance else "students"

    summary.value = (
        f"{tr['report_total']}: {total}  |  "
        f"{tr['report_covered']}: {covered_count}  |  "
        f"{tr.get('report_rescheduled', 'Rescheduled')}: {rescheduled_count}  |  "
        f"{tr['report_missed']}: {missed_count}\n"
        f"{tr.get('final_exam', 'Final Exam')}: {final_count} {final_unit}  |  "
        f"{tr.get('midterm_exam', 'Midterm Exam')}: {midterm_count} {midterm_unit}  |  "
        f"{tr.get('retake_exam', 'Retake Exam')}: {retake_count} {retake_unit}"
    )
    summary.font = Font(name="Arial", size=11, color="444444")
    summary.alignment = center

    filename = f"{teacher.full_name}_{month_name}_Report.xlsx".replace(" ", "_")
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.read())
    return response
def set_language(request):
    lang = request.GET.get("lang", "en")
    request.session["lang"] = lang
    next_url = request.GET.get("next", "/timetable/")
    return redirect(next_url)

from django.contrib.auth.models import User
from timetable.models import Department

@login_required(login_url="login")
def admin_panel(request):
    if not request.user.is_superuser:
        return redirect("timetable")

    lang = request.session.get("lang", "en")

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))
    num_days = calendar.monthrange(year, month)[1]
    prev_month = date(year, month, 1) - timedelta(days=1)
    next_month = date(year, month, num_days) + timedelta(days=1)

    teachers = Teacher.objects.all().select_related("user", "department").order_by("full_name")
    departments = Department.objects.all()

    total_teachers = teachers.count()
    with_accounts = teachers.filter(user__isnull=False).count()
    no_account = total_teachers - with_accounts

    # Fetch all submitted reports for this month, keep latest per teacher
    all_reports = MonthlyReport.objects.filter(
        year=year, month=month, status="submitted"
    ).order_by("teacher_id", "-id")

    latest_report_map = {}
    for r in all_reports:
        if r.teacher_id not in latest_report_map:
            latest_report_map[r.teacher_id] = r

    # Attach report object directly onto each teacher instance
    for t in teachers:
        t.latest_report = latest_report_map.get(t.id, None)

    total_submitted = len(latest_report_map)

    pending_dept_reports = DepartmentReport.objects.filter(
        year=year, month=month, status="submitted", approval_status="pending"
    ).count()

    return render(request, "admin_panel.html", {
        "teachers": teachers,
        "departments": departments,
        "total_teachers": total_teachers,
        "with_accounts": with_accounts,
        "no_account": no_account,
        "total_submitted": total_submitted,
        "pending_dept_reports": pending_dept_reports,
        "year": year,
        "month": month,
        "month_name": date(year, month, 1).strftime("%B %Y"),
        "prev": {"year": prev_month.year, "month": prev_month.month},
        "next": {"year": next_month.year, "month": next_month.month},
        "lang": lang,
    })


@login_required(login_url="login")
def admin_teacher_edit(request, teacher_id):
    if not request.user.is_superuser:
        return redirect("timetable")

    teacher = get_object_or_404(Teacher, id=teacher_id)
    departments = Department.objects.all()
    lang = request.session.get("lang", "en")
    message = None
    error = None

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()
        department_id = request.POST.get("department", "")
        is_dean = request.POST.get("is_dean") == "on"

        # Update department and dean status
        if department_id:
            teacher.department = Department.objects.filter(id=department_id).first()
        else:
            teacher.department = None
        teacher.is_dean = is_dean
        teacher.save()

        # Handle user account
        if username:
            if teacher.user:
                # Update existing user
                user = teacher.user
                if user.username != username:
                    if User.objects.filter(username=username).exclude(id=user.id).exists():
                        error = f"Username '{username}' is already taken."
                    else:
                        user.username = username
                if password:
                    user.set_password(password)
                if not error:
                    user.save()
                    message = "Teacher account updated successfully."
            else:
                # Create new user
                if User.objects.filter(username=username).exists():
                    error = f"Username '{username}' is already taken."
                else:
                    if not password:
                        error = "Password is required when creating a new account."
                    else:
                        user = User.objects.create_user(
                            username=username,
                            password=password
                        )
                        teacher.user = user
                        teacher.save()
                        message = "Teacher account created successfully."

        if not error and not message:
            message = "Teacher updated successfully."
    teacher.refresh_from_db()
    return render(request, "admin_teacher_edit.html", {
        "teacher": teacher,
        "departments": departments,
        "lang": lang,
        "message": message,
        "error": error,
    })

@login_required(login_url="login")
def admin_departments(request):
    if not request.user.is_superuser:
        return redirect("timetable")

    from timetable.models import Department
    lang = request.session.get("lang", "en")
    message = None
    error = None

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add":
            name = request.POST.get("name", "").strip()
            if not name:
                error = "Department name cannot be empty."
            elif Department.objects.filter(name=name).exists():
                error = f"Department '{name}' already exists."
            else:
                Department.objects.create(name=name)
                message = f"Department '{name}' added successfully."
        elif action == "delete":
            dept_id = request.POST.get("dept_id")
            dept = Department.objects.filter(id=dept_id).first()
            if dept:
                name = dept.name
                dept.delete()
                message = f"Department '{name}' deleted."

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    num_days = calendar.monthrange(year, month)[1]
    prev_month = date(year, month, 1) - timedelta(days=1)
    next_month = date(year, month, num_days) + timedelta(days=1)

    departments = Department.objects.all().order_by("name")

    dept_data = []
    for dept in departments:
        dept_teachers = dept.teachers.filter(user__isnull=False).order_by("-is_dean", "full_name")

        rows = []
        submitted_count = 0
        for t in dept_teachers:
            report = MonthlyReport.objects.filter(teacher=t, year=year, month=month).first()
            is_submitted = bool(report and report.status == "submitted")
            if is_submitted:
                submitted_count += 1
            rows.append({
                "teacher": t,
                "is_submitted": is_submitted,
                "report": report,
            })

        dept_report = DepartmentReport.objects.filter(
            department=dept, year=year, month=month
        ).first()

        dept_data.append({
            "department": dept,
            "rows": rows,
            "total_with_accounts": len(rows),
            "total_teachers": dept.teachers.count(),
            "submitted_count": submitted_count,
            "dept_report": dept_report,
        })
    month_name = f"{TRANSLATIONS[lang]['months'][month - 1]} {year}"
    return render(request, "admin_departments.html", {
        "dept_data": dept_data,
        "message": message,
        "error": error,
        "year": year,
        "month": month,
        "month_name": month_name,
        "prev": {"year": prev_month.year, "month": prev_month.month},
        "next": {"year": next_month.year, "month": next_month.month},
        "lang": lang,
    })

@login_required(login_url="login")
def admin_red_days(request):
    if not request.user.is_superuser:
        return redirect("timetable")

    from timetable.models import RedDay
    import calendar as cal_module
    import holidays as holidays_lib
    lang = request.session.get("lang", "en")
    message = None
    error = None

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add":
            date_str = request.POST.get("date", "").strip()
            reason = request.POST.get("reason", "").strip()
            if not date_str:
                error = "Date is required."
            elif not reason:
                error = "Reason is required."
            elif RedDay.objects.filter(date=date_str).exists():
                error = f"Red day already exists for {date_str}."
            else:
                RedDay.objects.create(
                    date=date_str,
                    reason=reason,
                    created_by=request.user
                )
                message = f"Red day added for {date_str}."
        elif action == "delete":
            red_day_id = request.POST.get("red_day_id")
            rd = RedDay.objects.filter(id=red_day_id).first()
            if rd:
                message = f"Red day {rd.date} removed."
                rd.delete()
        elif action == "delete_by_date":
            date_str = request.POST.get("date", "").strip()
            rd = RedDay.objects.filter(date=date_str).first()
            if rd:
                message = f"Red day {rd.date} removed."
                rd.delete()

    num_days = cal_module.monthrange(year, month)[1]
    first_weekday = cal_module.monthrange(year, month)[0]  # 0=Mon, 6=Sun

    # Get red days and holidays for this month
    red_days = RedDay.objects.filter(date__year=year, date__month=month).order_by("date")
    red_days_dict = {str(rd.date): rd.reason for rd in red_days}

    uz_holidays = holidays_lib.Uzbekistan(years=year)
    holidays_dict = {str(d): name for d, name in uz_holidays.items() if d.month == month}

    # Build calendar days
    calendar_days = []
    for day_num in range(1, num_days + 1):
        d = date(year, month, day_num)
        date_str = str(d)
        is_weekend = d.weekday() >= 6  # Sunday only (Saturday is working day at CIU)
        calendar_days.append({
            "day": day_num,
            "date": date_str,
            "is_red": date_str in red_days_dict,
            "reason": red_days_dict.get(date_str, ""),
            "is_holiday": date_str in holidays_dict,
            "holiday_name": holidays_dict.get(date_str, ""),
            "is_today": d == today,
            "is_weekend": is_weekend,
        })

    # Empty cells before first day (Monday=0)
    empty_days = range(first_weekday)

    prev_month = date(year, month, 1) - timedelta(days=1)
    next_month_date = date(year, month, num_days) + timedelta(days=1)
    month_name = f"{TRANSLATIONS[lang]['months'][month - 1]} {year}"
    return render(request, "admin_red_days.html", {
        "red_days": red_days,
        "calendar_days": calendar_days,
        "empty_days": empty_days,
        "message": message,
        "error": error,
        "year": year,
        "month": month,
        "month_name": month_name,
        "prev": {"year": prev_month.year, "month": prev_month.month},
        "next": {"year": next_month_date.year, "month": next_month_date.month},
        "lang": lang,
    })

def get_lesson_type(subject_name, classroom_name):
    subject_lower = (subject_name or "").lower()
    classroom_lower = (classroom_name or "").lower()

    # Check subject name first — specific exam types
    if any(k in subject_lower for k in ["final", "yakuniy"]):
        return "final_exam"
    if any(k in subject_lower for k in ["midterm", "oraliq", "mid-term"]):
        return "midterm_exam"
    if any(k in subject_lower for k in ["retake", "qayta", "repeat"]):
        return "retake_exam"
    if any(k in subject_lower for k in ["exam", "imtihon", "test"]):
        return "exam"

    # Only check room type if no exam keyword found in subject
    if "lecture" in classroom_lower or "ma'ruza" in classroom_lower:
        return "lecture"
    if "lab" in classroom_lower or "seminar" in classroom_lower:
        return "seminar"

    return "seminar"

@login_required(login_url="login")
@require_POST
def update_exam_count(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    if not teacher:
        return JsonResponse({"status": "error", "message": "No teacher linked"}, status=400)

    year = int(request.POST.get("year", date.today().year))
    month = int(request.POST.get("month", date.today().month))

    # Lock if report already submitted
    report = MonthlyReport.objects.filter(
        teacher=teacher, year=year, month=month
    ).order_by("-id").first()
    if (
            report
            and report.status == "submitted"
            and report.dean_approval != "rejected"
            and report.admin_approval != "rejected"
    ):
        return JsonResponse(
            {"status": "error", "message": "Report already submitted"},
            status=400,
        )

    field = request.POST.get("field", "")
    raw_value = request.POST.get("value", "0")
    unit = request.POST.get("unit", "students")

    valid_fields = {"final_exam_count", "midterm_exam_count", "retake_exam_count"}
    if field not in valid_fields:
        return JsonResponse({"status": "error", "message": "Invalid field"}, status=400)
    if unit not in ("groups", "students"):
        unit = "students"

    try:
        value = int(raw_value)
    except (ValueError, TypeError):
        return JsonResponse({"status": "error", "message": "Value must be a number"}, status=400)
    if value < 0:
        value = 0

    unit_field = field.replace("_count", "_unit")

    ea, _ = ExamAttendance.objects.get_or_create(
        teacher=teacher, year=year, month=month
    )
    setattr(ea, field, value)
    setattr(ea, unit_field, unit)
    ea.save(update_fields=[field, unit_field])

    return JsonResponse({"status": "ok", "field": field, "value": value, "unit": unit})

@login_required(login_url="login")
@require_POST
def add_lesson(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    if not teacher:
        return JsonResponse({"status": "error", "message": "No teacher account"})

    date_str = request.POST.get("date", "")
    period = request.POST.get("period", "")
    subject_name = request.POST.get("subject_name", "").strip()
    classroom_name = request.POST.get("classroom_name", "").strip()
    class_names = request.POST.get("class_names", "").strip()
    reason = request.POST.get("reason", "").strip()

    if not date_str or not period or not subject_name:
        return JsonResponse({"status": "error", "message": "Date, period and subject are required"})

    lesson_date = date.fromisoformat(date_str)
    day_name = DAYS[lesson_date.weekday()]

    # Find or create subject
    subject = Subject.objects.filter(name__iexact=subject_name).first()
    if not subject:
        subject = Subject.objects.create(
            edupage_id=f"custom_{teacher.id}_{subject_name[:20]}",
            name=subject_name
        )

    # Find classroom if provided
    classroom = None
    if classroom_name:
        classroom = Classroom.objects.filter(name__icontains=classroom_name).first()
        if not classroom:
            classroom = Classroom.objects.create(
                edupage_id=f"custom_{teacher.id}_{classroom_name[:30]}",
                name=classroom_name,
            )

    # Detect lesson type
    lesson_type = get_lesson_type(subject_name, classroom_name)

    # Check if slot already taken
    existing = ScheduledLesson.objects.filter(
        teacher=teacher,
        date=lesson_date,
        period=period
    ).first()

    if existing:
        return JsonResponse({"status": "error", "message": "This slot is already taken"})

    # Create the lesson
    sl = ScheduledLesson.objects.create(
        teacher=teacher,
        subject=subject,
        classroom=classroom,
        class_names=class_names,
        date=lesson_date,
        period=period,
        day=day_name,
        lesson_type=lesson_type,
        is_rescheduled_slot=False,
        is_custom_added=True,
    )
    if reason:
        LessonRecord.objects.create(
            teacher=teacher,
            scheduled_lesson=sl,
            date=lesson_date,
            is_covered=False,
            is_replaced=False,
            note=reason,
        )

    return JsonResponse({
        "status": "ok",
        "lesson_id": sl.id,
        "subject": str(sl.subject),
        "classroom": str(sl.classroom) if sl.classroom else "",
        "class_names": sl.class_names,
        "lesson_type": sl.lesson_type,
        "period": sl.period,
        "date": date_str,
    })


@login_required(login_url="login")
@require_POST
def delete_lesson(request, lesson_id):
    teacher = Teacher.objects.filter(user=request.user).first()
    sl = get_object_or_404(ScheduledLesson, id=lesson_id, teacher=teacher)

    # Only allow deleting manually added lessons or replacement slots
    # Never delete EduPage-synced lessons this way
    if sl.is_rescheduled_slot:
        # Replacement slot — also clear the original's record
        if sl.rescheduled_from:
            LessonRecord.objects.filter(
                scheduled_lesson=sl.rescheduled_from,
                is_replaced=True
            ).update(is_replaced=False, replacement_date=None)
        sl.delete()
        return JsonResponse({"status": "ok"})

    # Check if it was manually added (not from EduPage)
    # EduPage lessons have edupage_id starting with * on their subject
    # We mark manual ones by checking if subject edupage_id starts with "custom_"
    if sl.subject and sl.subject.edupage_id.startswith("custom_"):
        # Delete any lesson records too
        LessonRecord.objects.filter(scheduled_lesson=sl).delete()
        sl.delete()
        return JsonResponse({"status": "ok"})
    if sl.is_custom_added:
        LessonRecord.objects.filter(scheduled_lesson=sl).delete()
        sl.delete()
        return JsonResponse({"status": "ok"})

    return JsonResponse({"status": "error", "message": "Cannot delete EduPage lessons"})

@login_required(login_url="login")
@require_POST
def submit_department_report(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    if not teacher or not teacher.is_dean or not teacher.department:
        return JsonResponse({"status": "error", "message": "Not authorized"}, status=403)

    year = int(request.POST.get("year"))
    month = int(request.POST.get("month"))

    DepartmentReport.objects.update_or_create(
        department=teacher.department,
        year=year,
        month=month,
        defaults={
            "submitted_by": teacher,
            "submitted_at": timezone.now(),
            "status": "submitted",
            "approval_status": "pending",
            "approved_by": None,
            "approved_at": None,
            "admin_note": "",
        }
    )
    return JsonResponse({"status": "submitted"})


@login_required(login_url="login")
@require_POST
def cancel_department_report(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    if not teacher or not teacher.is_dean or not teacher.department:
        return JsonResponse({"status": "error", "message": "Not authorized"}, status=403)

    year = int(request.POST.get("year"))
    month = int(request.POST.get("month"))

    DepartmentReport.objects.filter(
        department=teacher.department,
        year=year,
        month=month,
        approval_status="pending",
    ).update(status="draft", submitted_at=None)

    return JsonResponse({"status": "cancelled"})


@login_required(login_url="login")
@require_POST
def review_department_report(request, report_id):

    if not request.user.is_superuser:
        return JsonResponse({"status": "error", "message": "Not authorized"}, status=403)

    decision = request.POST.get("decision")
    note = request.POST.get("note", "").strip()

    report = get_object_or_404(DepartmentReport, id=report_id)
    if report.status != "submitted":
        return JsonResponse({"status": "error", "message": "Report has not been submitted yet"}, status=400)
    if decision == "undo":
        if not request.user.is_superuser:
            return JsonResponse({"status": "error", "message": "Not authorized"}, status=403)
        report.approval_status = "pending"
        report.admin_note = ""
        report.approved_by = None
        report.approved_at = None
        report.save()
        return JsonResponse({"status": "ok"})
    if decision == "approve":
        report.approval_status = "approved"
    elif decision == "reject":
        report.approval_status = "rejected"
    else:
        return JsonResponse({"status": "error", "message": "Invalid decision"}, status=400)

    report.approved_by = request.user
    report.approved_at = timezone.now()
    report.admin_note = note
    report.save()

    return JsonResponse({"status": "ok", "approval_status": report.approval_status})


@login_required(login_url="login")
@require_POST
def review_teacher_report(request, report_id):
    report = get_object_or_404(MonthlyReport, id=report_id)
    decision = request.POST.get("decision")
    note = request.POST.get("note", "").strip()
    requesting_teacher = Teacher.objects.filter(user=request.user).first()
    is_dean_of_report = (
            requesting_teacher and requesting_teacher.is_dean
            and requesting_teacher.department_id == report.teacher.department_id
    )
    if decision == "undo":
        actor = request.POST.get("actor", "")
        if actor == "dean":
            if not is_dean_of_report:
                return JsonResponse({"status": "error", "message": "Not authorized"}, status=403)
            report.dean_approval = "pending"
            report.dean_rejection_reason = ""
            report.admin_approval = "waiting_dean"  # reset admin track too
            report.save()
        elif actor == "admin":
            if not request.user.is_superuser:
                return JsonResponse({"status": "error", "message": "Not authorized"}, status=403)
            report.admin_approval = "pending"
            report.admin_rejection_reason = ""
            report.save()
        else:
            return JsonResponse({"status": "error", "message": "Invalid actor"}, status=400)
        return JsonResponse({"status": "ok"})

    if decision not in ("approve", "reject"):
        return JsonResponse({"status": "error", "message": "Invalid decision"}, status=400)
    if report.status != "submitted":
        return JsonResponse({"status": "error", "message": "Report has not been submitted yet"}, status=400)



    if report.dean_approval == "pending":
        if not is_dean_of_report:
            return JsonResponse({"status": "error", "message": "Only the department dean can review this report"}, status=403)

        report.dean_reviewed_by = requesting_teacher
        report.dean_reviewed_at = timezone.now()
        if decision == "approve":
            report.dean_approval = "approved"
            report.admin_approval = "pending"
        else:
            report.dean_approval = "rejected"
            report.dean_rejection_reason = note
        report.save()
        return JsonResponse({"status": "ok", "dean_approval": report.dean_approval, "admin_approval": report.admin_approval})

    elif report.dean_approval == "approved" and report.admin_approval == "pending":
        if not request.user.is_superuser:
            return JsonResponse({"status": "error", "message": "Only an admin can review this report"}, status=403)

        dept_report = DepartmentReport.objects.filter(
            department=report.teacher.department, year=report.year, month=report.month, status="submitted"
        ).first()
        if not dept_report:
            return JsonResponse({"status": "error", "message": "Department report has not been submitted yet"}, status=403)

        report.admin_reviewed_by = request.user
        report.admin_reviewed_at = timezone.now()
        if decision == "approve":
            report.admin_approval = "approved"
        else:
            # Admin rejection resets dean approval too — teacher must resubmit from scratch
            report.admin_approval = "rejected"
            report.admin_rejection_reason = note
            report.dean_approval = "pending"
            report.dean_rejection_reason = ""
        report.save()
        return JsonResponse({"status": "ok", "dean_approval": report.dean_approval, "admin_approval": report.admin_approval})

    else:
        return JsonResponse({"status": "error", "message": "This report cannot be reviewed at this stage"}, status=400)


@login_required(login_url="login")
@require_POST
def bulk_review_teacher_reports(request, dept_id):
    decision = request.POST.get("decision")
    year = int(request.POST.get("year"))
    month = int(request.POST.get("month"))
    scope = request.POST.get("scope", "submitted")

    if decision not in ("approve", "reject"):
        return JsonResponse({"status": "error", "message": "Invalid decision"}, status=400)

    from timetable.models import Department
    department = get_object_or_404(Department, id=dept_id)
    requesting_teacher = Teacher.objects.filter(user=request.user).first()
    is_dean_of_dept = (
        requesting_teacher and requesting_teacher.is_dean
        and requesting_teacher.department_id == department.id
    )

    if request.user.is_superuser:
        dept_report = DepartmentReport.objects.filter(department=department, year=year, month=month, status="submitted").first()
        if not dept_report:
            return JsonResponse({"status": "error", "message": "Department report has not been submitted yet"}, status=403)
        stage_field, stage_value = "admin_approval", "pending"
    elif is_dean_of_dept:
        stage_field, stage_value = "dean_approval", "pending"
    else:
        return JsonResponse({"status": "error", "message": "Not authorized"}, status=403)

    dept_teacher_ids = list(Teacher.objects.filter(department=department, user__isnull=False).values_list("id", flat=True))

    latest_ids = []
    for tid in dept_teacher_ids:
        r = MonthlyReport.objects.filter(teacher_id=tid, year=year, month=month).order_by("-id").first()
        if r and r.status == "submitted" and getattr(r, stage_field) == stage_value:
            latest_ids.append(r.id)

    if scope == "all" and len(latest_ids) < len(dept_teacher_ids):
        return JsonResponse({"status": "error", "message": f"Not all teachers have a report waiting at this stage ({len(latest_ids)} of {len(dept_teacher_ids)})."}, status=400)
    if not latest_ids:
        return JsonResponse({"status": "error", "message": "No reports waiting for review at this stage"}, status=400)

    reports = MonthlyReport.objects.filter(id__in=latest_ids)
    now = timezone.now()

    if stage_field == "dean_approval":
        if decision == "approve":
            reports.update(dean_approval="approved", admin_approval="pending", dean_reviewed_by=requesting_teacher, dean_reviewed_at=now)
        else:
            reports.update(dean_approval="rejected", dean_reviewed_by=requesting_teacher, dean_reviewed_at=now, dean_rejection_reason="")
    else:
        if decision == "approve":
            reports.update(admin_approval="approved", admin_reviewed_by=request.user, admin_reviewed_at=now)
        else:
            reports.update(admin_approval="rejected", admin_reviewed_by=request.user, admin_reviewed_at=now, admin_rejection_reason="")

    return JsonResponse({"status": "ok", "updated": len(latest_ids)})

@login_required(login_url="login")
def view_teacher_report(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)

    requesting_teacher = Teacher.objects.filter(user=request.user).first()
    is_self = requesting_teacher and requesting_teacher.id == teacher.id
    is_their_dean = (
        requesting_teacher and requesting_teacher.is_dean
        and requesting_teacher.department_id == teacher.department_id
    )
    if not (request.user.is_superuser or is_self or is_their_dean):
        return redirect("timetable")

    lang = request.session.get("lang", "en")
    tr = TRANSLATIONS[lang]
    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))
    month_name = f"{tr['months'][month - 1]} {year}"

    prev_month = date(year, month, 1) - timedelta(days=1)
    next_month = (date(year, month, 1) + timedelta(days=32)).replace(day=1)

    prev = {"year": prev_month.year, "month": prev_month.month}
    next_ = {"year": next_month.year, "month": next_month.month}

    if request.user.is_superuser:
        back_url = f"/department-report/{teacher.department_id}/?year={year}&month={month}" if teacher.department_id else "/admin-panel/departments/"
    elif is_their_dean:
        back_url = "/dean/"
    else:
        back_url = "/timetable/"

    # GATE: admin cannot see until dean approved AND dept report submitted
    if request.user.is_superuser:
        latest = MonthlyReport.objects.filter(
            teacher=teacher, year=year, month=month
        ).order_by("-id").first()

        dean_approved = latest and latest.dean_approval == "approved"
        dept_report_check = DepartmentReport.objects.filter(
            department=teacher.department,
            year=year,
            month=month,
            status="submitted"
        ).first() if teacher.department else None

        if not dean_approved or not dept_report_check:
            return render(request, "teacher_report_view.html", {
                "teacher": teacher,
                "not_visible_yet": True,
                "year": year,
                "month": month,
                "month_name": month_name,
                "prev": prev,
                "next": next_,
                "back_url": back_url,
                "lang": lang,
                "tr": tr,
            })

    # Build report data
    history = MonthlyReport.objects.filter(
        teacher=teacher, year=year, month=month
    ).order_by("id")
    report = history.last()
    is_submitted = report is not None and report.status == "submitted"

    admin_can_review = False
    if request.user.is_superuser and report and report.dean_approval == "approved" and report.admin_approval == "pending":
        dept_report = DepartmentReport.objects.filter(
            department=teacher.department, year=year, month=month, status="submitted"
        ).first()
        admin_can_review = bool(dept_report)

    lessons = ScheduledLesson.objects.filter(
        teacher=teacher, date__year=year, date__month=month,
        is_rescheduled_slot=False,
    ).select_related("subject", "classroom").order_by("date", "period")

    records = LessonRecord.objects.filter(teacher=teacher, date__year=year, date__month=month)
    record_map = {r.scheduled_lesson_id: r for r in records}

    rows = []
    covered_count = 0
    rescheduled_count = 0
    missed_count = 0

    for lesson in lessons:
        record = record_map.get(lesson.id)

        if record and record.is_replaced:
            if record.is_covered:
                status, status_class = tr["report_covered"], "covered"
            elif is_submitted:
                status, status_class = tr["report_missed"], "missed"
            else:
                status, status_class = tr["report_pending"], "rescheduled"

            if record.replacement_date:
                rep_date_str = record.replacement_date.strftime("%d %b %Y")
                rep_slot = ScheduledLesson.objects.filter(
                    teacher=teacher, date=record.replacement_date, is_rescheduled_slot=True
                ).first()
                note = f"Rescheduled → {rep_date_str} {tr['report_period']} {rep_slot.period}" if rep_slot else f"Rescheduled → {rep_date_str}"
                if record.note:
                    note += f" — {record.note}"
            else:
                note = "Rescheduled"
                if record.note:
                    note += f" — {record.note}"

            rescheduled_count += 1
            if status_class == "covered":
                covered_count += 1
            elif status_class == "missed":
                missed_count += 1

        elif record and record.is_covered:
            status, status_class = tr["report_covered"], "covered"
            note = record.note or "—"
            covered_count += 1

        elif record and record.is_covered == False:
            if is_submitted:
                status, status_class = tr["report_missed"], "missed"
                missed_count += 1
            else:
                status, status_class = tr["report_pending"], "pending"
            note = record.note or "—"

        else:
            if is_submitted:
                status, status_class = tr["report_missed"], "missed"
                missed_count += 1
            else:
                status, status_class = tr["report_pending"], "pending"
            note = "—"

        if hasattr(lesson, 'is_custom_added') and lesson.is_custom_added:
            marker = tr.get("manually_added", "Manually added by teacher")
            note = f"{marker} — {note}" if note and note != "—" else marker

        type_labels = {
            "lecture": tr.get("type_lecture", "Lecture"),
            "seminar": tr.get("type_seminar", "Seminar"),
            "lab": tr.get("type_lab", "Lab"),
            "exam": tr.get("type_exam", "Exam"),
            "final_exam": tr.get("type_final_exam", "Final Exam"),
            "midterm_exam": tr.get("type_midterm_exam", "Midterm Exam"),
            "retake_exam": tr.get("type_retake_exam", "Retake Exam"),
        }

        rows.append({
            "date": lesson.date,
            "day": tr["days_full"][lesson.date.weekday()],
            "subject": str(lesson.subject) if lesson.subject else "—",
            "period": lesson.period,
            "classroom": str(lesson.classroom) if lesson.classroom else "—",
            "type_label": type_labels.get(lesson.lesson_type, lesson.lesson_type.capitalize()),
            "status": status,
            "status_class": status_class,
            "note": note,
        })

    return render(request, "teacher_report_view.html", {
        "teacher": teacher,
        "rows": rows,
        "report": report,
        "history": history,
        "is_submitted": is_submitted,
        "is_dean_viewer": is_their_dean,
        "is_admin": request.user.is_superuser,
        "admin_can_review": admin_can_review,
        "total": len(rows),
        "covered_count": covered_count,
        "rescheduled_count": rescheduled_count,
        "missed_count": missed_count,
        "year": year,
        "month": month,
        "month_name": month_name,
        "prev": prev,
        "next": next_,
        "back_url": back_url,
        "lang": lang,
        "tr": tr,
    })
    # ... rest of the view continues unchanged ...

@login_required(login_url="login")
def view_department_report(request, dept_id):
    from timetable.models import Department
    department = get_object_or_404(Department, id=dept_id)

    requesting_teacher = Teacher.objects.filter(user=request.user).first()
    is_their_dean = requesting_teacher and requesting_teacher.is_dean and requesting_teacher.department_id == department.id
    if not (request.user.is_superuser or is_their_dean):
        return redirect("timetable")

    lang = request.session.get("lang", "en")
    tr = TRANSLATIONS[lang]

    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))
    month_name = f"{tr['months'][month - 1]} {year}"
    num_days = calendar.monthrange(year, month)[1]
    all_days = [date(year, month, d) for d in range(1, num_days + 1)]
    working_days = [d for d in all_days if d.weekday() < 6]
    holidays = get_uz_holidays(year, month)
    prev_month = date(year, month, 1) - timedelta(days=1)
    next_month = date(year, month, num_days) + timedelta(days=1)
    dept_report = DepartmentReport.objects.filter(
        department=department, year=year, month=month
    ).first()

    dept_report_submitted = bool(dept_report and dept_report.status == "submitted")

    if request.user.is_superuser and not dept_report_submitted:
        return render(request, "department_report_view.html", {
            "department": department,
            "not_submitted_yet": True,
            "year": year, "month": month, "month_name": month_name,
            "prev": {"year": prev_month.year, "month": prev_month.month},
            "next": {"year": next_month.year, "month": next_month.month},
            "lang": lang, "tr": tr, "is_admin": True,

        })
    dept_teachers = Teacher.objects.filter(
        department=department, user__isnull=False
    ).order_by("-is_dean", "full_name")

    teacher_rows = []
    for t in dept_teachers:
        records = LessonRecord.objects.filter(
            teacher=t, date__year=year, date__month=month, is_covered=True
        )
        day_counts = {str(d): records.filter(date=d).count() for d in working_days}
        total = records.count()
        report = MonthlyReport.objects.filter(teacher=t, year=year, month=month).order_by("-id").first()
        ea = ExamAttendance.objects.filter(teacher=t, year=year, month=month).first()
        teacher_rows.append({
            "teacher": t,
            "day_counts": day_counts,
            "total": total,

            "report": report,
            "final_exam_count": ea.final_exam_count if ea else 0,
        "final_exam_unit": ea.final_exam_unit if ea else "students",
        "midterm_exam_count": ea.midterm_exam_count if ea else 0,
        "midterm_exam_unit": ea.midterm_exam_unit if ea else "students",
        "retake_exam_count": ea.retake_exam_count if ea else 0,
        "retake_exam_unit": ea.retake_exam_unit if ea else "students",
            # "is_submitted": bool(report and report.status == "submitted"),
        })





    return render(request, "department_report_view.html", {
        "department": department,
        "teacher_rows": teacher_rows,
        "working_days": working_days,
        "holidays": holidays,
        "dept_report": dept_report,
        "year": year,
        "month": month,
        "month_name": month_name,
        "prev": {"year": prev_month.year, "month": prev_month.month},
        "next": {"year": next_month.year, "month": next_month.month},
        "lang": lang,
        "tr": tr,
        "is_admin": request.user.is_superuser,
        "week_spans": get_week_spans(working_days),
    })

def get_period_list():
    """Get periods from DB, fall back to hardcoded if DB empty"""
    from timetable.models import Period as PeriodModel
    db_periods = PeriodModel.objects.all().order_by("number")
    if db_periods.exists():
        return [
            {"number": p.number, "start": p.start_time, "end": p.end_time}
            for p in db_periods
        ]
    # Fallback to hardcoded
    return [
        {"number": p, "start": PERIOD_TIMES[p][0], "end": PERIOD_TIMES[p][1]}
        for p in PERIODS
    ]

def get_periods_list_numbers():
    """Get just period numbers"""
    from timetable.models import Period as PeriodModel
    db_periods = PeriodModel.objects.all().order_by("number")
    if db_periods.exists():
        return [p.number for p in db_periods]
    return PERIODS

@login_required(login_url="login")
def preview_report(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    if not teacher:
        return redirect("timetable")

    lang = request.session.get("lang", "en")
    tr = TRANSLATIONS[lang]
    today = date.today()
    year = int(request.GET.get("year", today.year))
    month = int(request.GET.get("month", today.month))
    month_name = f"{tr['months'][month - 1]} {year}"

    num_days = calendar.monthrange(year, month)[1]
    all_days = [date(year, month, d) for d in range(1, num_days + 1)]
    working_days = [d for d in all_days if d.weekday() < 6]
    holidays = get_uz_holidays(year, month)

    records_covered = LessonRecord.objects.filter(
        teacher=teacher, date__year=year, date__month=month, is_covered=True
    )
    day_counts = {str(d): records_covered.filter(date=d).count() for d in working_days}
    total = records_covered.count()

    report = MonthlyReport.objects.filter(
        teacher=teacher, year=year, month=month
    ).order_by("-id").first()
    already_submitted = report and report.status == "submitted"

    # Build daily report rows (same logic as view_teacher_report)
    lessons = ScheduledLesson.objects.filter(
        teacher=teacher, date__year=year, date__month=month,
        is_rescheduled_slot=False,
    ).select_related("subject", "classroom").order_by("date", "period")

    all_records = LessonRecord.objects.filter(teacher=teacher, date__year=year, date__month=month)
    record_map = {r.scheduled_lesson_id: r for r in all_records}

    rows = []
    covered_count = 0
    rescheduled_count = 0
    missed_count = 0

    type_labels = {
        "lecture": tr.get("type_lecture", "Lecture"),
        "seminar": tr.get("type_seminar", "Seminar"),
        "lab": tr.get("type_lab", "Lab"),
        "exam": tr.get("type_exam", "Exam"),
        "final_exam": tr.get("type_final_exam", "Final Exam"),
        "midterm_exam": tr.get("type_midterm_exam", "Midterm Exam"),
        "retake_exam": tr.get("type_retake_exam", "Retake Exam"),
    }

    for lesson in lessons:
        record = record_map.get(lesson.id)

        if record and record.is_replaced:
            if record.is_covered:
                status, status_class = tr["report_covered"], "covered"
                covered_count += 1
            else:
                status, status_class = tr["report_pending"], "rescheduled"
            rescheduled_count += 1
            if record.replacement_date:
                rep_date_str = record.replacement_date.strftime("%d %b %Y")
                rep_slot = ScheduledLesson.objects.filter(
                    teacher=teacher, date=record.replacement_date, is_rescheduled_slot=True
                ).first()
                note = f"Rescheduled → {rep_date_str} {tr['report_period']} {rep_slot.period}" if rep_slot else f"Rescheduled → {rep_date_str}"
                if record.note:
                    note += f" — {record.note}"
            else:
                note = "Rescheduled"
                if record.note:
                    note += f" — {record.note}"

        elif record and record.is_covered:
            status, status_class = tr["report_covered"], "covered"
            note = record.note or "—"
            covered_count += 1

        else:
            status, status_class = tr["report_pending"], "pending"
            note = record.note if record and record.note else "—"

        if hasattr(lesson, 'is_custom_added') and lesson.is_custom_added:
            marker = tr.get("manually_added", "Manually added by teacher")
            note = f"{marker} — {note}" if note and note != "—" else marker

        rows.append({
            "date": lesson.date,
            "day": tr["days_full"][lesson.date.weekday()],
            "subject": str(lesson.subject) if lesson.subject else "—",
            "period": lesson.period,
            "classroom": str(lesson.classroom) if lesson.classroom else "—",
            "type_label": type_labels.get(lesson.lesson_type, lesson.lesson_type.capitalize()),
            "status": status,
            "status_class": status_class,
            "note": note,
        })

    prev_month = date(year, month, 1) - timedelta(days=1)
    next_month = date(year, month, num_days) + timedelta(days=1)
    exam_attendance = ExamAttendance.objects.filter(
        teacher=teacher, year=year, month=month
    ).first()
    return render(request, "report_preview.html", {
        "teacher": teacher,
        "working_days": working_days,
        "holidays": holidays,
        "day_counts": day_counts,
        "total": total,
        "rows": rows,
        "covered_count": covered_count,
        "rescheduled_count": rescheduled_count,
        "missed_count": missed_count,
        "report": report,
        "already_submitted": already_submitted,
        "year": year,
        "month": month,
        "month_name": month_name,
        "prev": {"year": prev_month.year, "month": prev_month.month},
        "next": {"year": next_month.year, "month": next_month.month},
        "lang": lang,
        "tr": tr,
        "week_spans": get_week_spans(working_days),
        "exam_attendance": exam_attendance,
    })

def get_week_spans(working_days):
    """Returns list of (week_number, day_count) for colspan headers."""
    spans = []
    week_num = 1
    count = 0
    for i, d in enumerate(working_days):
        if i == 0:
            count = 1
        elif d.weekday() == 0:  # Monday starts a new week
            spans.append((week_num, count))
            week_num += 1
            count = 1
        else:
            count += 1
    if count > 0:
        spans.append((week_num, count))
    return spans

@login_required(login_url="login")
@require_POST
def mark_lessons_bulk(request):
    teacher = Teacher.objects.filter(user=request.user).first()
    if not teacher:
        return JsonResponse({"status": "error"}, status=400)

    lesson_ids = request.POST.get("lesson_ids", "").split(",")
    dates = request.POST.get("dates", "").split(",")
    is_covered = request.POST.get("is_covered", "true") == "true"

    if len(lesson_ids) != len(dates):
        return JsonResponse({"status": "error", "message": "Mismatched ids/dates"}, status=400)

    results = []
    from django.db import transaction
    with transaction.atomic():
        for lesson_id, date_str in zip(lesson_ids, dates):
            lesson_id = lesson_id.strip()
            date_str = date_str.strip()
            if not lesson_id or not date_str:
                continue
            try:
                sl = ScheduledLesson.objects.get(id=lesson_id, teacher=teacher)
                lesson_date = date.fromisoformat(date_str)
                existing = LessonRecord.objects.filter(
                    scheduled_lesson=sl, date=lesson_date
                ).first()
                record, _ = LessonRecord.objects.update_or_create(
                    scheduled_lesson=sl,
                    date=lesson_date,
                    defaults={
                        "teacher": teacher,
                        "is_covered": is_covered,
                        "note": existing.note if existing else "",
                        "is_replaced": existing.is_replaced if existing else False,
                        "replacement_date": existing.replacement_date if existing else None,
                    }
                )
                results.append({
                    "lesson_id": lesson_id,
                    "date": date_str,
                    "is_covered": record.is_covered,
                    "is_replaced": record.is_replaced,
                })
            except (ScheduledLesson.DoesNotExist, ValueError):
                continue

    return JsonResponse({"status": "ok", "results": results})